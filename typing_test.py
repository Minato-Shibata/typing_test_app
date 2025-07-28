import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
import random
import time
import json
import os
import matplotlib.pyplot as plt
import japanize_matplotlib
from matplotlib.ticker import MaxNLocator
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from datetime import datetime
import statistics

SCORE_FILE = "scores.json"

def load_sentences_from_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    sentences = [cell.value for cell in ws['A'] if cell.value]
    return sentences

def save_score(char_count):
    scores = load_scores()
    scores.append({"timestamp": datetime.now().isoformat(), "correct_chars": char_count})
    with open(SCORE_FILE, "w") as f:
        json.dump(scores, f, indent=2)

def load_scores():
    if not os.path.exists(SCORE_FILE):
        return []
    with open(SCORE_FILE, "r") as f:
        return json.load(f)

def clear_scores():
    if os.path.exists(SCORE_FILE):
        os.remove(SCORE_FILE)

class TypingTestApp:
    def __init__(self, root, sentences):
        self.root = root
        self.all_sentences = sentences
        self.timer_label = tk.Label(root, text="経過時間: 00:00", font=("Arial", 10), anchor='e')
        self.timer_label.pack(anchor='se', padx=10, pady=5)
        self.show_title()

    def reset_state(self):
        self.index = 0
        self.correct_count = 0
        self.total_correct_chars = 0
        self.results = []
        self.start_time = None
        self.timer_running = False
        self.sentences = self.all_sentences[:]
        random.shuffle(self.sentences)

    def show_title(self):
        self.clear_widgets()
        self.title_label = tk.Label(self.root, text="タイピングテスト", font=("Arial", 24))
        self.title_label.pack(pady=40)

        start_btn = tk.Button(self.root, text="スタート", font=("Arial", 14), command=self.start_test)
        score_btn = tk.Button(self.root, text="スコア確認", font=("Arial", 14), command=self.show_score_graph)
        start_btn.pack(pady=10)
        score_btn.pack(pady=10)

    def start_test(self, event=None):
        self.clear_widgets()
        self.reset_state()
        self.setup_ui()
        self.start_time = time.time()
        self.timer_running = True
        self.update_timer()
        self.next_sentence()
        self.root.after(300000, self.force_finish)  # 5分

    def setup_ui(self):
        self.label = tk.Label(self.root, text="", font=("Arial", 16))
        self.label.pack(pady=20)

        self.entry = tk.Entry(self.root, font=("Arial", 14), width=80)
        self.entry.pack(pady=10)
        self.entry.bind("<Return>", self.check_answer)
        self.entry.focus()

        self.feedback = tk.Label(self.root, text="", font=("Arial", 12))
        self.feedback.pack()

    def update_timer(self):
        if self.start_time and self.timer_running:
            elapsed = int(time.time() - self.start_time)
            minutes = elapsed // 60
            seconds = elapsed % 60
            self.timer_label.config(text=f"経過時間: {minutes:02}:{seconds:02}")
            self.root.after(1000, self.update_timer)

    def next_sentence(self):
        if self.index < len(self.sentences):
            self.current_sentence = self.sentences[self.index]
            self.label.config(text=f"問題 {self.index + 1}: {self.current_sentence}")
            self.entry.delete(0, tk.END)
        else:
            self.show_results()

    def check_answer(self, event=None):
        user_input = self.entry.get().strip()
        correct = user_input == self.current_sentence.strip()
        self.results.append({
            "問題": self.current_sentence,
            "回答": user_input,
            "正誤": "正解" if correct else "不正解"
        })

        if correct:
            self.correct_count += 1
            self.total_correct_chars += len(self.current_sentence.strip())

        self.index += 1
        self.next_sentence()

    def force_finish(self):
        if self.index < len(self.sentences):
            self.timer_running = False
            self.timer_label.config(text="経過時間: 00:00")
            self.label.config(text="⏰ 時間切れ！")
            self.entry.destroy()
            self.feedback.destroy()
            self.show_results()

    def show_results(self):
        self.timer_running = False
        self.timer_label.config(text="経過時間: 00:00")
        self.clear_widgets()
        save_score(self.total_correct_chars)

        result_text = f"✅ 正解数: {self.correct_count} / {len(self.results)}\n"
        result_text += f"📝 正解した文章の合計文字数: {self.total_correct_chars}\n\n"
        result_text += "Enterキーを押すとタイトル画面に戻ります。\n\n"

        self.text_box = tk.Text(self.root, wrap=tk.WORD, font=("Arial", 12))
        self.text_box.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        self.text_box.tag_config("incorrect_text", foreground="red")

        self.text_box.insert(tk.END, result_text)

        for idx, r in enumerate(self.results, start=1):
            block = f"【問題 {idx}】{r['問題']}\n【回答】{r['回答']}\n【結果】{r['正誤']}\n\n"
            start_idx = self.text_box.index(tk.END)
            self.text_box.insert(tk.END, block)
            end_idx = self.text_box.index(tk.END)

            if r["正誤"] == "不正解":
                self.text_box.tag_add("incorrect_text", start_idx, end_idx)

        self.text_box.config(state=tk.DISABLED)
        self.root.bind("<Return>", self.return_to_title)

    def return_to_title(self, event=None):
        self.root.unbind("<Return>")
        self.text_box.destroy()
        self.show_title()

    def show_score_graph(self):
        self.clear_widgets()
        scores = load_scores()

        if not scores:
            label = tk.Label(self.root, text="スコアデータがありません。", font=("Arial", 14))
            label.pack(pady=20)
            back_btn = tk.Button(self.root, text="戻る", command=self.show_title)
            back_btn.pack(pady=10)
            return

        self.scores = scores
        self.graph_type = tk.StringVar(value="折れ線グラフ")

        # 統計情報を表示
        values = [s["correct_chars"] for s in scores]
        max_val = max(values)
        min_val = min(values)
        avg_val = round(statistics.mean(values), 2)

        stats_frame = tk.Frame(self.root)
        stats_frame.pack(pady=10)
        tk.Label(stats_frame, text=f"最高点: {max_val}").grid(row=0, column=0, padx=5)
        tk.Label(stats_frame, text=f"最低点: {min_val}").grid(row=0, column=1, padx=5)
        tk.Label(stats_frame, text=f"平均点: {avg_val}").grid(row=0, column=2, padx=5)

        # グラフタイプ選択
        option_frame = tk.Frame(self.root)
        option_frame.pack(pady=5)
        ttk.Combobox(option_frame, textvariable=self.graph_type,
                     values=["折れ線グラフ", "箱ひげ図"], state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(option_frame, text="表示更新", command=self.update_graph).pack(side=tk.LEFT, padx=5)

        # グラフ描画領域
        self.graph_area = tk.Frame(self.root)
        self.graph_area.pack(pady=10)
        self.update_graph()

        # 戻る・リセットボタン
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=20)
        back_btn = tk.Button(btn_frame, text="戻る", command=self.show_title)
        reset_btn = tk.Button(btn_frame, text="スコアをすべて削除", command=self.confirm_clear_scores)
        back_btn.pack(side=tk.LEFT, padx=10)
        reset_btn.pack(side=tk.LEFT, padx=10)

    def update_graph(self):
        for w in self.graph_area.winfo_children():
            w.destroy()

        values = [s["correct_chars"] for s in self.scores]
        fig, ax = plt.subplots(figsize=(6, 4))

        if self.graph_type.get() == "折れ線グラフ":
            x = range(1, len(values) + 1)
            ax.plot(x, values, marker='o')
            ax.set_title("スコア推移")
            ax.set_xlabel("回数")
            ax.set_ylabel("正解文字数")
            ax.xaxis.set_major_locator(MaxNLocator(integer=True))
            for i, n in enumerate(values):
                ax.text(x[i], values[i], n)
        else:
            ax.boxplot(values)
            ax.set_title("スコア分布（箱ひげ図）")
            ax.set_ylabel("正解文字数")

        canvas = FigureCanvasTkAgg(fig, master=self.graph_area)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def confirm_clear_scores(self):
        if messagebox.askyesno("確認", "本当にスコアを削除しますか？"):
            clear_scores()
            messagebox.showinfo("完了", "スコアを削除しました。")
            self.show_score_graph()

    def clear_widgets(self):
        for widget in self.root.winfo_children():
            if widget != self.timer_label:
                widget.destroy()

# メイン実行
if __name__ == "__main__":
    filepath = "typing_test.xlsx"
    sentences = load_sentences_from_excel(filepath)

    root = tk.Tk()
    root.title("タイピングテスト")
    root.geometry("800x600")
    app = TypingTestApp(root, sentences)
    root.mainloop()
